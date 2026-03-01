#!/usr/bin/env python3
"""
WEEB-bot: search Nyaa.si for next file numbers in series and add magnet links to qBittorrent.
Requirements:
    pip install openpyxl qbittorrent-api requests beautifulsoup4 nyaapy
(If nyaapy isn't available, the script will fall back to scraping nyaa.si.)
"""

import re
import time
import base64
import random
import requests
from datetime import datetime, timezone
from types import ModuleType
from typing import Optional, Dict, Any, List, Tuple, cast
from openpyxl import load_workbook
from bs4 import BeautifulSoup, Tag

try:
    from nyaapy import Nyaa
    NYAAPI_AVAILABLE = True
except Exception:
    NYAAPI_AVAILABLE = False

qbittorrentapi: Optional[ModuleType] = None
try:
    import qbittorrentapi
except Exception:
    pass


# ============================================================
#                   USER CONFIGURATION
# ============================================================

# --- Excel file settings ---
EXCEL_PATH       = "Downloadlist.xlsx"       # Path to your Excel file
TAB_PARAMS       = "Parameters_Series"       # Sheet name for series parameters
TAB_LOG          = "Download_Log"            # Sheet name for the download log
TAB_CREDENTIALS  = "QB_Credentials"          # Sheet name for qBittorrent credentials
                                             # See the QB_Credentials sheet in Downloadlist.xlsx
                                             # to set your qBittorrent username, password and category.
                                             # These are intentionally NOT stored here to avoid
                                             # exposing sensitive information in the repository.

# --- qBittorrent Web UI connection ---
QB_HOST = "http://127.0.0.1"                 # Host where qBittorrent Web UI is running
QB_PORT = 8080                               # Port for qBittorrent Web UI
                                             # Username, password and category are loaded from
                                             # the QB_Credentials sheet in the Excel file.

# --- Episode number matching ---
EPISODE_NUMBER_SEARCH_WINDOW = 7             # Number of characters after the series name to search for the episode number.
                                             # Increase if your titles have longer separators between title and episode number.
                                             # Example: "Series Name - 11" needs at least 5 (" - 11"). "Series Name Episode 11" would need more.

# --- Episode catch-up limit ---
MAX_EPISODES_PER_SERIES = 12                 # Maximum number of episodes to download per series per run.
                                             # Set to None for unlimited (will keep going until no more are found).

# --- Nyaa search settings ---
NYAA_SEARCH_CATEGORY  = "1_2"                        # Nyaa category code to search in. "1_2" = Anime - English-translated.
NYAA_CATEGORY_FILTER  = "anime - english-translated" # The category string to validate against in candidate filtering.
                                                     # Must match what Nyaa's HTML returns for the above category code.

# --- Scraping politeness delays (in seconds) ---
SCRAPE_DELAY_MIN      = 2.5                  # Minimum delay before each Nyaa scrape request
SCRAPE_DELAY_MAX      = 5.0                  # Maximum delay before each Nyaa scrape request

# --- Inter-episode delay (in seconds) ---
# Delay between checking consecutive episodes of the same series in the catch-up loop
INTER_EPISODE_DELAY_MIN = 3.0
INTER_EPISODE_DELAY_MAX = 5.0

# --- qBittorrent confirmation settings ---
QB_CONFIRM_TIMEOUT    = 8.0                  # How many seconds to wait for qBittorrent to confirm a torrent was added
QB_CONFIRM_POLL       = 0.5                  # How often (in seconds) to poll qBittorrent during confirmation wait

# ============================================================
#               END OF USER CONFIGURATION
# ============================================================


# ---------- Helper ----------
def _strip_value(val: Any) -> Any:
    """Strip whitespace from a cell value if it is a string; leave other types untouched."""
    if isinstance(val, str):
        return val.strip()
    return val


# ---------- Credentials loader ----------
def load_qb_credentials(path: str) -> Dict[str, str]:
    """
    Load qBittorrent credentials from the QB_Credentials sheet in the Excel file.
    Expected columns: QB_Username, QB_Password, QB_Category
    Values are read from row 2 (the first data row below the header).
    Raises RuntimeError if the sheet is missing or any required field is empty.
    """
    wb = load_workbook(path)
    if TAB_CREDENTIALS not in wb.sheetnames:
        raise RuntimeError(
            f"Credentials sheet '{TAB_CREDENTIALS}' not found in '{path}'.\n"
            f"Please add a sheet named '{TAB_CREDENTIALS}' with columns: QB_Username, QB_Password, QB_Category.\n"
            f"See the README for instructions."
        )
    ws   = wb[TAB_CREDENTIALS]
    rows = list(ws.rows)
    if len(rows) < 2:
        raise RuntimeError(
            f"Credentials sheet '{TAB_CREDENTIALS}' has no data row. "
            f"Row 1 must be the header (QB_Username, QB_Password, QB_Category) and row 2 must contain your values."
        )

    # Strip whitespace from both headers and values to prevent invisible trailing
    # spaces in Excel cells from causing silent authentication failures.
    headers = [_strip_value(c.value) for c in rows[0]]
    data    = {headers[i]: _strip_value(rows[1][i].value if i < len(rows[1]) else None)
               for i in range(len(headers))}

    username = data.get("QB_Username")
    password = data.get("QB_Password")
    category = data.get("QB_Category")

    missing = [name for name, val in [("QB_Username", username), ("QB_Password", password), ("QB_Category", category)] if not val]
    if missing:
        raise RuntimeError(
            f"The following required fields are empty in the '{TAB_CREDENTIALS}' sheet: {', '.join(missing)}.\n"
            f"Please fill them in before running WEEB-bot."
        )

    return {
        "qb_username": str(username),
        "qb_password": str(password),
        "qb_category": str(category),
    }


def load_download_log_latest(path: str) -> Dict[str, int]:
    wb = load_workbook(path)
    if TAB_LOG not in wb.sheetnames:
        return {}
    ws   = wb[TAB_LOG]
    rows = list(ws.rows)
    if len(rows) < 1:
        return {}

    # Strip whitespace from headers to prevent invisible spaces in Excel
    # from causing column lookups to silently return None.
    headers = [_strip_value(c.value) for c in rows[0]]
    mapping: Dict[str, int] = {}

    for r in rows[1:]:
        if all(cell.value is None for cell in r):
            continue
        row     = {headers[i]: (r[i].value if i < len(r) else None) for i in range(len(headers))}
        name    = row.get("SeriesName")
        filenum = row.get("FileNumber")
        try:
            filenum = int(filenum) if filenum is not None and filenum != "" else None
        except (TypeError, ValueError):
            filenum = None
        if name and filenum is not None:
            if name not in mapping or (filenum > mapping[name]):
                mapping[name] = filenum
    return mapping


def load_series_parameters(path: str) -> List[Dict[str, Any]]:
    """Load enabled series from Excel and build their search query for next file number."""
    wb = load_workbook(path)
    if TAB_PARAMS not in wb.sheetnames:
        raise RuntimeError(f"Workbook missing required sheet '{TAB_PARAMS}'")
    ws   = wb[TAB_PARAMS]
    rows = list(ws.rows)
    if len(rows) < 1:
        return []

    headers     = [_strip_value(c.value) for c in rows[0]]
    series_list = []
    log_latest  = load_download_log_latest(path)

    for idx, r in enumerate(rows[1:], start=2):
        if all(cell.value is None for cell in r):
            continue
        row     = {headers[i]: (r[i].value if i < len(r) else None) for i in range(len(headers))}
        enabled = str(row.get("Enabled", "Yes")).strip().lower()
        if enabled in ("no", "false", "0", "n"):
            continue

        raw_num = row.get("CurrentFileNumber")
        try:
            current_num = int(raw_num) if raw_num is not None and raw_num != "" else None
        except (TypeError, ValueError):
            current_num = None

        series_name_key = row.get("SeriesName") or ""
        cur_num         = current_num if current_num is not None else log_latest.get(series_name_key, 0)
        target_num      = int(cur_num) + 1

        parts        = [row.get(col) for col in ["Publisher", "SeriesName", "Resolution", "OtherFilters"]]
        parts        = [str(p).strip() for p in parts if p not in (None, "")]
        search_query = " ".join(parts + [str(target_num)])

        series_list.append({
            "SeriesName":        str(row.get("SeriesName") or ""),
            "Publisher":         str(row.get("Publisher") or ""),
            "Resolution":        str(row.get("Resolution") or ""),
            "OtherFilters":      str(row.get("OtherFilters") or ""),
            "CurrentFileNumber": current_num,
            "TargetNumber":      target_num,
            "Query":             search_query,
            "RowIndex":          idx,
        })

    return series_list


# ---------- Candidate Filtering ----------
def episode_distance_in_title(title: str, series_name: str, target_number: int) -> int:
    """
    Return the character distance between the first occurrence of the series name
    and the episode number in the title. Smaller distance is better.
    """
    if not title or not series_name:
        return 9999

    t_lower     = title.lower()
    s_lower     = series_name.lower()
    num_pattern = re.compile(rf"(?<!\d)0{{0,3}}{target_number}(?!\d)")

    s_idx = t_lower.find(s_lower)
    if s_idx == -1:
        return 9999

    m = num_pattern.search(t_lower)
    if not m:
        return 9999

    return abs(m.start() - s_idx)


def debug_candidate_rejection(title: str, category: str, series_name: str,
                              target_number: int, resolution: str, publisher: str) -> Tuple[bool, str, int]:
    if not title or not series_name:
        return False, "missing title or series name", 9999

    title_lower      = title.lower()
    series_lower     = series_name.lower()
    resolution_lower = resolution.lower() if resolution else ""
    publisher_lower  = publisher.lower() if publisher else ""

    # 1. Category check
    if NYAA_CATEGORY_FILTER not in category.lower():
        return False, f"category '{category}' != '{NYAA_CATEGORY_FILTER}'", 9999

    # 2. Series name check
    series_idx = title_lower.find(series_lower)
    if series_idx == -1:
        return False, f"series name '{series_name}' not in title", 9999
    series_end_idx = series_idx + len(series_lower)

    # 3. Episode number check
    # Extract the N characters immediately following the end of the series name
    # and check whether the target episode number appears there.
    # N is set by EPISODE_NUMBER_SEARCH_WINDOW in the user configuration.
    search_window = title_lower[series_end_idx : series_end_idx + EPISODE_NUMBER_SEARCH_WINDOW]
    ep_pattern    = rf"(?<!\d)0{{0,3}}{target_number}(?!\d)"
    if not re.search(ep_pattern, search_window):
        return False, (
            f"episode number {target_number} not found in the {EPISODE_NUMBER_SEARCH_WINDOW} "
            f"characters after series name (window: '{search_window}')"
        ), 9999

    # 4. Resolution check
    if resolution_lower and resolution_lower not in title_lower:
        return False, f"resolution '{resolution_lower}' not found in title", 9999

    # 5. Publisher check
    if publisher_lower and publisher_lower not in title_lower:
        return False, f"publisher '{publisher_lower}' not found in title", 9999

    dist = episode_distance_in_title(title, series_name, target_number)
    return True, f"passed all filters (distance={dist})", dist


# ---------- Nyaa search ----------
def search_nyaa_by_scrape(query: str, category: Optional[str] = None) -> List[Dict[str, Any]]:
    base   = "https://nyaa.si"
    params: Dict[str, str] = {"f": "0", "q": query}
    if category:
        params["c"] = category

    try:
        delay = random.uniform(SCRAPE_DELAY_MIN, SCRAPE_DELAY_MAX)
        print(f"  Delaying {delay:.1f}s before scraping Nyaa...")
        time.sleep(delay)
        r = requests.get(base + "/", params=params, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"  Failed to fetch Nyaa.si page: {e}")
        return []

    soup      = BeautifulSoup(r.text, "html.parser")
    table_raw = soup.find("table", {"class": "torrent-list"}) or soup.find("tbody")

    results: List[Dict[str, Any]] = []

    if not isinstance(table_raw, Tag):
        print("  WARNING: No torrent table found in page HTML.")
        return results

    table = table_raw

    for tr_elem in table.find_all("tr"):
        tr = cast(Tag, tr_elem)
        if 'class' in tr.attrs and 'header' in (tr.get("class") or []):
            continue
        tds_raw = tr.find_all("td")
        if not tds_raw or len(tds_raw) < 2:
            continue

        tds = [cast(Tag, td) for td in tds_raw]

        a_tag = None
        for a in tds[1].find_all("a", href=True):
            a = cast(Tag, a)
            if "comments" not in (a.get("class") or []) and not str(a.get("href", "")).endswith("#comments"):
                a_tag = a
                break
        title_text = a_tag.text.strip() if a_tag else "Unknown Title"

        magnet_link = None
        for a in (tds[2].find_all("a", href=True) if len(tds) > 2 else []):
            a = cast(Tag, a)
            if str(a.get("href", "")).startswith("magnet:"):
                magnet_link = str(a["href"])
                break

        date_text   = tds[4].text.strip() if len(tds) > 4 else None
        cat_tag_raw = tds[0].find("a", title=True)
        cat_tag     = cat_tag_raw if isinstance(cat_tag_raw, Tag) else None
        category_text = str(cat_tag["title"]).strip() if cat_tag and "title" in cat_tag.attrs else "Unknown"

        results.append({
            "title":        title_text,
            "release_date": date_text,
            "magnet":       magnet_link,
            "category":     category_text
        })

    return results


def search_nyaa(query: str, category: Optional[str] = None) -> List[Dict[str, Any]]:
    if NYAAPI_AVAILABLE:
        try:
            n       = Nyaa()
            res     = n.search(query)
            results = []
            for r in res:
                results.append({
                    "title":        getattr(r, "title", str(r)),
                    "release_date": getattr(r, "date", None) or getattr(r, "pub_date", None),
                    "magnet":       getattr(r, "magnet", None) or getattr(r, "magnet_link", None),
                    "category":     getattr(r, "category", "Unknown")
                })
            if results:
                return results
        except Exception as e:
            print(f"  nyaapy search failed for query '{query}': {e}")

    try:
        return search_nyaa_by_scrape(query, category)
    except Exception as e:
        print(f"  Scrape search failed for query '{query}': {e}")
        return []


# ---------- qBittorrent integration ----------
def get_qb_client(config: Dict[str, Any]) -> Any:
    if qbittorrentapi is None:
        raise RuntimeError("qbittorrent-api is not installed. Run: pip install qbittorrent-api")
    host   = config["qb_host"].rstrip("/")
    port   = config["qb_port"]
    client = qbittorrentapi.Client(
        host=f"{host}:{port}",
        username=config["qb_username"],
        password=config["qb_password"]
    )
    try:
        client.auth_log_in()
    except qbittorrentapi.LoginFailed as e:
        raise RuntimeError(f"Could not log into qBittorrent Web UI at {host}:{port}: {e}")
    return client


def extract_infohash_from_magnet(magnet: str) -> Optional[str]:
    if not magnet:
        return None
    m = re.search(r'xt=urn:btih:([A-Fa-f0-9]{40}|[A-Za-z2-7]{32,40})', magnet)
    if not m:
        return None
    h = m.group(1)
    if len(h) == 40 and re.fullmatch(r'[A-Fa-f0-9]{40}', h, flags=re.IGNORECASE):
        return h.lower()
    try:
        b32     = h.upper()
        padding = '=' * ((8 - len(b32) % 8) % 8)
        b       = base64.b32decode(b32 + padding)
        return b.hex()
    except Exception:
        return h.lower()


def add_magnet_and_confirm(client: Any, magnet_link: str, category: Optional[str] = None) -> bool:
    try:
        kwargs: Dict[str, Any] = {"urls": magnet_link}
        if category:
            kwargs["category"] = category
        client.torrents_add(**kwargs)
    except Exception as e:
        print(f"  qB API add failed: {e}")
        return False

    infohash = extract_infohash_from_magnet(magnet_link)
    timeout  = time.time() + QB_CONFIRM_TIMEOUT

    if not infohash:
        while time.time() < timeout:
            try:
                for t in client.torrents_info():
                    mu = getattr(t, 'magnet_uri', None) or getattr(t, 'magnet_link', None) or getattr(t, 'magnet', None)
                    if mu and magnet_link.split('&')[0] in mu:
                        return True
            except Exception:
                pass
            time.sleep(QB_CONFIRM_POLL)
        return False

    while time.time() < timeout:
        try:
            if client.torrents_info(hashes=infohash):
                return True
        except Exception:
            pass
        time.sleep(QB_CONFIRM_POLL)
    return False


# ---------- Excel updates ----------
def log_and_update_excel(path: str, entry: Dict[str, Any], series_name: str, new_number: int) -> None:
    """
    Write a log entry to Download_Log AND update CurrentFileNumber in Parameters_Series
    in a single workbook open/save operation.
    """
    wb = load_workbook(path)

    # --- Write to Download_Log ---
    if TAB_LOG not in wb.sheetnames:
        ws_log = wb.create_sheet(TAB_LOG)
        ws_log.append(["SeriesName", "FileNumber", "DownloadedFileName", "ReleaseDate",
                        "DownloadDate", "MagnetLink", "Status", "Notes"])
    else:
        ws_log = wb[TAB_LOG]

    ws_log.append([
        entry.get("SeriesName"),
        entry.get("FileNumber"),
        entry.get("DownloadedFileName"),
        entry.get("ReleaseDate"),
        entry.get("DownloadDate"),
        entry.get("MagnetLink"),
        entry.get("Status"),
        entry.get("Notes", ""),
    ])

    # --- Update CurrentFileNumber in Parameters_Series ---
    if TAB_PARAMS in wb.sheetnames:
        ws_params = wb[TAB_PARAMS]
        headers   = [_strip_value(c.value) for c in list(ws_params.rows)[0]]
        try:
            idx_name = headers.index("SeriesName")
            idx_cur  = headers.index("CurrentFileNumber")
            found    = False
            for row in ws_params.iter_rows(min_row=2):
                if row[idx_name].value == series_name:
                    row[idx_cur].value = new_number
                    found = True
                    break
            if not found:
                print(f"  WARNING: Could not find '{series_name}' in '{TAB_PARAMS}' to update CurrentFileNumber.")
        except ValueError:
            print(f"  WARNING: Could not find 'SeriesName' or 'CurrentFileNumber' column in '{TAB_PARAMS}'.")
    else:
        print(f"  WARNING: Sheet '{TAB_PARAMS}' not found when trying to update CurrentFileNumber.")

    wb.save(path)


# ---------- Main ----------
def main() -> None:
    print("Loading credentials from", EXCEL_PATH)
    try:
        credentials = load_qb_credentials(EXCEL_PATH)
    except RuntimeError as e:
        print(f"ERROR: {e}")
        return

    config: Dict[str, Any] = {
        "qb_host":     QB_HOST,
        "qb_port":     QB_PORT,
        "qb_username": credentials["qb_username"],
        "qb_password": credentials["qb_password"],
        "qb_category": credentials["qb_category"],
    }

    print("Connecting to qBittorrent Web UI...")
    try:
        qb_client = get_qb_client(config)
        print("Connected to qBittorrent Web UI.")
    except Exception as e:
        print(f"ERROR: Could not connect to qBittorrent Web UI: {e}")
        print("Make sure qBittorrent is running and the Web UI is enabled before running WEEB-bot.")
        return

    print("Loading series from", EXCEL_PATH)
    series_list = load_series_parameters(EXCEL_PATH)
    if not series_list:
        print("No enabled series found in the Parameters_Series sheet.")
        return

    for s in series_list:
        name           = s["SeriesName"]
        publisher      = s["Publisher"]
        resolution     = s["Resolution"]
        other          = s["OtherFilters"]
        target_num     = s["TargetNumber"]
        episodes_added = 0

        if not name:
            print("Skipping a row with no SeriesName.")
            continue

        print(f"\n--- Processing series: '{name}' ---")

        while True:
            if MAX_EPISODES_PER_SERIES is not None and episodes_added >= MAX_EPISODES_PER_SERIES:
                print(f"  Reached the episode cap of {MAX_EPISODES_PER_SERIES} for '{name}'. Moving to next series.")
                break

            query_parts = [p for p in [publisher, name, resolution, other] if p not in ("", None)]
            query       = " ".join(query_parts + [str(target_num)])

            print(f"\n  Searching for '{name}' episode {target_num} (query: {query})")
            try:
                results = search_nyaa(query, category=NYAA_SEARCH_CATEGORY)
            except Exception as e:
                print(f"  Search failed for '{name}': {e}")
                break

            if not results:
                print("  No results found. Moving to next series.")
                break

            print(f"  {len(results)} result(s) found:")
            for idx, r in enumerate(results, start=1):
                print(f"    {idx}. {r.get('title')} | Category: {r.get('category')} | Date: {r.get('release_date')}")

            valid_candidates: List[Tuple[int, Dict[str, Any]]] = []
            for r in results:
                if not r.get("magnet"):
                    print(f"  Skipping '{r.get('title')}' -> no magnet link.")
                    continue

                valid, reason, dist = debug_candidate_rejection(
                    r.get("title") or "", r.get("category") or "",
                    name, target_num, resolution, publisher
                )
                if valid:
                    print(f"  Accepted: '{r.get('title')}' -> {reason}")
                    valid_candidates.append((dist, r))
                else:
                    print(f"  Rejected: '{r.get('title')}' -> {reason}")

            if not valid_candidates:
                print(f"  No valid candidate found for '{name}' #{target_num}. Moving to next series.")
                break

            valid_candidates.sort(key=lambda x: x[0])
            chosen = valid_candidates[0][1]
            print(f"  Final choice: '{chosen.get('title')}' (distance={valid_candidates[0][0]})")

            print("  Adding torrent to qBittorrent...")
            accepted = add_magnet_and_confirm(qb_client, chosen["magnet"], category=config["qb_category"])
            if not accepted:
                print(f"  qBittorrent did not confirm the torrent for '{name}' #{target_num}. Will retry next run.")
                break

            print("  Updating download log and series list...")
            log_and_update_excel(
                EXCEL_PATH,
                entry={
                    "SeriesName":         name,
                    "FileNumber":         target_num,
                    "DownloadedFileName": chosen.get("title"),
                    "ReleaseDate":        chosen.get("release_date"),
                    "DownloadDate":       datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                    "MagnetLink":         chosen.get("magnet"),
                    "Status":             "Added",
                    "Notes":              ""
                },
                series_name=name,
                new_number=target_num
            )
            print(f"  Successfully queued and logged '{name}' #{target_num}.")

            episodes_added += 1
            target_num     += 1

            delay = random.uniform(INTER_EPISODE_DELAY_MIN, INTER_EPISODE_DELAY_MAX)
            print(f"  Waiting {delay:.1f}s before checking next episode...")
            time.sleep(delay)

    print("\nLogging out of qBittorrent Web UI...")
    try:
        qb_client.auth_log_out()
    except Exception:
        pass

    print("Done.")


if __name__ == "__main__":
    main()
